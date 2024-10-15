
from contextlib import redirect_stderr
from django.shortcuts import render,redirect
from django.contrib.auth import login,logout
from urllib3 import HTTPResponse
from product.models import Product
from django.views import View
from .forms import *
from openpyxl import Workbook,load_workbook

from django.contrib.auth import login,authenticate,logout
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.http import JsonResponse
from django.views import View
from django.shortcuts import get_object_or_404
from .models import Product
from django.contrib.auth.models import User
from django.conf import settings




# Create your views here.



class Home(View):
    def get(self, request):
        return render(request, 'home.html')

class Register(View):
    def get(self, request):
        return render(request, 'register.html')


    def post(self, request):
        
        if request.method == 'POST':
            username = request.POST.get('username')
            password = request.POST.get('password')
            confirm_password = request.POST.get('confirm_password')
            
            if password != confirm_password:
                error = 'Passwords do not match'
                return render(request, 'register.html', {'error': error})
            
            try:
                user = User.objects.create_user(username=username, password=password)
                user.save()
                login(request, user)
                return redirect('login')
            except Exception as e:
                error = 'Username already exists'
                return render(request, 'register.html', {'error': error})
        else:
            return HTTPResponse('Invalid')
                  
class Login(View):
    def get(self,request):
        return render(request,'login.html')
    
    def post(self,request):
      
        if request.method == 'POST':
            username=request.POST.get('username')
            password=request.POST.get('password')
            print(username,password)
            user=authenticate(request,username=username,password=password)
            print(user)
            if user is not None:
                login(request,user)
                return redirect('home')
            else:
                error='Invalid username or password'
                return render(request,'login.html',{'error':error})
        else:
            return HTTPResponse('Invalid')
        
class Logout(LoginRequiredMixin,View):
    login_url='login'
    def get(self, request):
        logout(request)
        return redirect('login')  

class ProductCreate(View):
    def get(self, request):
        # Check if the user is the specified user
        if request.user.username in settings.ALLOWED_USERS:
            form = ProductForm()
            return render(request, 'product_create.html', {'form': form})
        return redirect('product_list')
    
    def post(self, request):
        if request.user.username == 'sohith':
            form = ProductForm(request.POST)
            if form.is_valid():
                form.save()
                return redirect('product_list')
            return render(request, 'product_create.html', {'form': form})
        return redirect('product_list')

class ProductUpdate(View):
    def get(self, request, id):
        if request.user.username in settings.ALLOWED_USERS:
            product_instance = get_object_or_404(Product, id=id)
            form = ProductForm(instance=product_instance)
            return render(request, 'product_update.html', {'form': form})
        return redirect('product_list')
    
    def post(self, request, id):
        if request.user.username in settings.ALLOWED_USERS:
            product_instance = get_object_or_404(Product, id=id)
            form = ProductForm(request.POST, instance=product_instance)
            if form.is_valid():
                form.save()
                return redirect('product_list')
            return render(request, 'product_update.html', {'form': form})
        return redirect('product_list')


class ProductDelete(View):
    def get(self, request, id):
        if request.user.username in settings.ALLOWED_USERS:
            product_instance = get_object_or_404(Product, id=id)
            product_instance.delete()
        return redirect('product_list')

class ProductList(View):
    def get(self,request):
        products=Product.objects.all()
        allowed_users= settings.ALLOWED_USERS 
        return render(request, 'product_list.html', {'products': products,'allowed_users': allowed_users})
    
    def post(self, request):
        product_id = request.POST.get('product_id')
        stock_to_add = request.POST.get('stock_to_add')

        if product_id and stock_to_add and stock_to_add.isdigit():
            product_instance = get_object_or_404(Product, id=product_id)
            product_instance.stock += int(stock_to_add)
            product_instance.save()

        return redirect('product_list')
    
    


class ExportProducts(View):

    def get(self, request, *args, **kwargs):
       
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Data"
        
      
        headers = ['ID', 'Name', 'Price', 'Stock', 'Status']
        ws.append(headers)  
        
        products = Product.objects.all()
        for product in products:
            ws.append([product.id, product.name, product.price, product.stock, product.get_status_display()])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
        
        wb.save(response)
        return response


  

class ImportProducts(View):
    def get(self, request, *args, **kwargs):
        form = ExcelForm()  
        return render(request, 'import_products.html', {'form': form})

    def post(self, request, *args, **kwargs):
        form = ExcelForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            wb = load_workbook(file)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    
                    product_id, name, price, stock, status = row[:5]  

                    product, created = Product.objects.update_or_create(
                        id=product_id,
                        defaults={
                            'name': name,
                            'price': price,
                            'stock': stock,
                            'status': status if status in dict(Product.STATUS_CHOICES) else Product.AVAILABLE
                        }
                    )
                except ValueError as e:
                    print(f"Skipping row: {row}. Error: {e}")

            return redirect('product_list')  
        return render(request, 'import_products.html', {'form': form})




class OrderCreateView(LoginRequiredMixin, View):
    def get(self, request):
        product_id = request.GET.get('product_id')
        if not product_id:
            return redirect('product_list')
        product = get_object_or_404(Product, id=product_id)
        form = OrderForm(initial={'product': product})
        return render(request, 'order/order_form.html', {'form': form, 'product': product})

    def post(self, request):
        form = OrderForm(request.POST)
        if form.is_valid():
            quantity = form.cleaned_data['quantity']
            product_id = form.cleaned_data['product'].id
            product = get_object_or_404(Product, id=product_id)

            if quantity > product.stock:
                return render(request, 'order/order_form.html', {'form': form, 'product': product, 'error': "The requested quantity exceeds available stock."})

            order = form.save(commit=False)
            order.user = request.user  # Assign the logged-in user to the order
            order.save()
            return redirect('order_list')

        return render(request, 'order/order_form.html', {'form': form})


# Read Order List View
class OrderListView(LoginRequiredMixin, View):
    def get(self, request):
        orders = Order.objects.filter(user=request.user)  # Retrieve orders for the logged-in user
        return render(request, 'order/order_list.html', {'orders': orders})



# Update Order View
class OrderUpdateView(View):
    def get(self, request, pk):
        order = get_object_or_404(Order, pk=pk)
        form = OrderForm(instance=order)
        return render(request, 'order/order_form.html', {'form': form, 'order': order})

    def post(self, request, pk):
        order = get_object_or_404(Order, pk=pk)
        form = OrderForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            return redirect('order_list')  # Redirect to order list after successful update
        return render(request, 'order/order_form.html', {'form': form, 'order': order})

# Delete Order View
class OrderDeleteView(View):
    def get(self, request, pk):
        order = get_object_or_404(Order, pk=pk)
        product = get_object_or_404(Product, name=order.product)
        product.stock += order.quantity
        product.save()
        order.delete()
        return redirect('order_list')  # Redirect to order list after deletion
    
class OrderDetailView(LoginRequiredMixin, View):
    def get(self, request, pk):
        order = get_object_or_404(Order, pk=pk, user=request.user)  # Filter by user
        return render(request, 'order/order_details.html', {'order': order})
